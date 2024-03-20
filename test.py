from customtkinter import *
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from main import *

set_appearance_mode("dark")
set_default_color_theme("green")

app = CTk()
app.geometry("650x500")
app.title("Schedule Maker")

def make_excel_automate():

    wb2 = load_workbook("sheets/sheet.xlsx")
    ws2 = wb2.active
    
    wb = load_workbook("Template/תבנית זמינות.xlsx")
    ws = wb.active

    wb3 = load_workbook("Template/תבנית סידור.xlsx")
    ws3 = wb3.active

    date_str = datetime.now().strftime("%Y-%m-%d")
    week_date = datetime.strptime(date_str, "%Y-%m-%d")
    
    
    # C = date, D - name, E - Sunday ...., K=Saturday, L=notes, M=email
    workers = []
    curr_row_in_template = 1
    for i in range(ws2.max_row,2,-1):
        curr_date_str = timestamp_to_date(ws2["C"+str(i)].value)
        curr_date = datetime.strptime(curr_date_str, "%Y-%m-%d")
        # checking if the date on the table is the same as this week date.
        if curr_date.isocalendar()[1] == week_date.isocalendar()[1]:
            if ws2["M"+str(i)].value not in workers:
                curr_row_in_template+=1
                workers.append(ws2["M"+str(i)].value)
                # logic to insert the demands into the template
                for j in range(8):
                    # A represent name in template, D represent name in sheet
                    ws[chr(65 + j)+str(curr_row_in_template)] = ws2[chr(68 + j)+str(i)].value
                ws3['J'+str(curr_row_in_template)] =  ws[chr(65)+str(curr_row_in_template)].value      
    
    # get the backup worker to be at the end of the workers list
    ws3['J'+str(curr_row_in_template+1)] = ws3['H'+str(16)].value
    ws3['J'+str(curr_row_in_template+1)].fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')
    ws3['H'+str(16)].fill = PatternFill(fill_type=None)
    ws3['H'+str(16)] = None
    
    # saving the excel files.
    wb.save("automate/זמינות.xlsx")
    wb3.save("automate/תבנית סידור.xlsx")
    main(manual=False)
    



''' This function convert timestamp cell to date'''
def timestamp_to_date(timestamp_str):
    
    # getting the date without the ending in the brackets
    timestamp_str = timestamp_str.split("(")[0].strip()

    #format of timestamp
    timestamp_format = "%a %b %d %Y %H:%M:%S gmt%z"

    #string to datetime object
    timestamp_dt = datetime.strptime(timestamp_str, timestamp_format)

    # datetime object to a regular date string
    regular_date_str = timestamp_dt.strftime("%Y-%m-%d")

    return regular_date_str


def manual():
    main()




frame = CTkFrame(master=app)
frame.pack(pady=20,padx=30, fill="both", expand=True)

label = CTkLabel(master=frame,text="Make a schedule with google-forms")
label.pack(pady=12,padx=10)

textbox = CTkTextbox(master=frame, width=20,height=60)
textbox.insert("0.0", "1. Please enter the sheet file in the folder sheets and save the updated file \
                        \n2. Press on the \" automate\" button.")
textbox.pack(pady=20,padx=20,fill="both",expand=False)

button = CTkButton(master=frame, text="AUTOMATE", command=make_excel_automate)
button.pack(pady=12,padx=10)


textbox = CTkTextbox(master=frame, width=20,height=60)
textbox.insert("0.0", "I fill manually all the workers and the settings I want in the excel file: \"זמינות\" \nand now I can click on manual")
textbox.pack(pady=20,padx=20,fill="both",expand=False)


button = CTkButton(master=frame, text="MANUAL",command=manual)
button.pack(pady=12,padx=10)




app.mainloop()