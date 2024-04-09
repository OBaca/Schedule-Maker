from customtkinter import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from data.algorithm import algorithm
from data.constants import *
from data.init_table import *

def gui():
    set_appearance_mode("dark")
    set_default_color_theme("green")

    app = CTk()
    app.geometry("650x500")
    app.title("Schedule Maker")

    frame = CTkFrame(master=app)
    frame.pack(pady=20,padx=30, fill="both", expand=True)

    label = CTkLabel(master=frame,text="Make a schedule with google-forms")
    label.pack(pady=12,padx=10)

    textbox = CTkTextbox(master=frame, width=20,height=60)
    textbox.insert("0.0", "1. Please enter the sheet file in the folder sheets and save the updated file \
                            \n2. Press on the \" automate\" button.")
    textbox.pack(pady=20,padx=20,fill="both",expand=False)

    button = CTkButton(master=frame, text="AUTOMATE", command=make_excel_automate)
    button.pack(pady=12,padx=10)


    button = CTkButton(master=frame, text="CREATE NEW TABLE",command=create_new_table)
    button.pack(pady=12,padx=10)
    
    textbox = CTkTextbox(master=frame, width=20,height=20)
    textbox.insert("0.0", "CREATE NEW TABLE")
    textbox.pack(pady=20,padx=20,fill="both",expand=False)

    entry_name = CTkEntry(frame,placeholder_text="Enter schedule name (in english)",width=220,corner_radius=50,fg_color=("green","lightgreen"),placeholder_text_color="grey",text_color="black")
    entry_name.pack(pady=20)

    entry_amount_of_workers = CTkEntry(frame,placeholder_text="Enter amount of workers",width=170,corner_radius=50,fg_color=("green","lightgreen"),placeholder_text_color="grey",text_color="black")
    entry_amount_of_workers.pack(pady=20)

    entry_amount_of_stations = CTkEntry(frame,placeholder_text="Enter amount of stations",width=170,corner_radius=50,fg_color=("green","lightgreen"),placeholder_text_color="grey",text_color="black")
    entry_amount_of_stations.pack(pady=20)

    button = CTkButton(master=frame, text="CREATE NEW TABLE",fg_color=['#3B8ED0','#1F6AA5'],command=lambda : create_new_table(entry_name.get(),int(entry_amount_of_workers.get()),int(entry_amount_of_stations.get())))
    button.pack(pady=12,padx=10)
    

    app.mainloop()

'''
def create_new_table(name,amount_workers,amount_stations):
    print(f"Name:{name},amount:{amount_workers},stations:{amount_stations}")
    wb = load_workbook("create_new_table/new_table.xlsx")
    ws = wb.active
    try:
        ws['B2'] = name
        ws['B3'] = int(amount_workers)
        ws['B4'] = int(amount_stations)
    except:
        print("Please use english only on name, please use whole numbers on amount")


    wb.save("create_new_table/new_table.xlsx")
'''

def make_excel_automate():

    wb2 = load_workbook("sheets/sheet.xlsx")
    ws2 = wb2.active
    
    wb = load_workbook("Template/תבנית זמינות.xlsx")
    ws = wb.active

    wb3 = load_workbook("Template/תבנית סידור.xlsx")
    ws3 = wb3.active

    date_str = datetime.now().strftime("%Y-%m-%d")
    week_date = datetime.strptime(date_str, "%Y-%m-%d")
    

    comment_position_on_sheets = 'J'
    
    # C = date, D - name, E - Sunday ...., K=Saturday, L=notes, M=email
    workers = []
    curr_row_in_template = 1
    for i in range(ws2.max_row,1,-1):
        print(curr_row_in_template)
        if not ws2["A"+str(i)].value:
            continue
        curr_date_str = timestamp_to_date(ws2["A"+str(i)].value)
        curr_date = datetime.strptime(curr_date_str, "%Y-%m-%d")
        # checking if the date on the table is the same as this week date.
        if curr_date.isocalendar()[1] == week_date.isocalendar()[1]:
            if ws2["K"+str(i)].value not in workers:
                curr_row_in_template+=1
                workers.append(ws2["K"+str(i)].value)
                # logic to insert the demands into the template
                for j in range(8):
                    # A represent name in template, D represent name in sheet
                    ws[chr(65 + j)+str(curr_row_in_template)] = ws2[chr(66 + j)+str(i)].value

                ws3['J'+str(curr_row_in_template)] =  ws[chr(65)+str(curr_row_in_template)].value
                ws[COMMENT_POS_ON_CONSTRAINTS+str(curr_row_in_template)] = ws2[comment_position_on_sheets+str(i)].value
    
    # get the backup worker to be at the end of the workers list
    ws3['J'+str(curr_row_in_template+1)] = ws3['H'+str(16)].value
    ws3['J'+str(curr_row_in_template+1)].fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')
    ws3['H'+str(16)].fill = PatternFill(fill_type=None)
    ws3['H'+str(16)] = None
    


    # saving the excel files.
    wb.save("automate/תבנית זמינות.xlsx")
    wb3.save("automate/תבנית סידור.xlsx")

    wb.close()
    wb2.close()
    wb3.close()
    algorithm()
    



''' This function convert timestamp cell to date'''
def timestamp_to_date(timestamp_str):
    
    # getting the date without the ending in the brackets
    timestamp_str = timestamp_str.split("(")[0].strip()

    #format of timestamp
    timestamp_format = "%a %b %d %Y %H:%M:%S gmt%z"

    #string to datetime object
    timestamp_dt = datetime.strptime(timestamp_str, timestamp_format)

    # datetime object to a regular date string
    regular_date_str = timestamp_dt.strftime("%Y-%m-%d")

    return regular_date_str


